import base64
import json
from core import *
from core.invoice import Invoice
from core.log import LogLevel
from core.invoice.baidu_ocr import BaiduOCR
from core.utils import extract_params_from_url, extract_text
from sqlite_utils import Database
from tqdm import tqdm
from yaspin import yaspin
import custom_rule

UPLOADER_COLUMN_NAME = "创建人"
BELONGER_COLUMN_NAME = "收款人"
INVOICE_COLUMN_NAME = "发票"
TOTAL_AMOUNT_COLUMN_NAME = "审批后金额"
APPROVAL_REMARKS_COLUMN_NAME = "审批备注"

LARK_LOG_LEVEL = LogLevel.INFO

table_fields_type_map = {
    "file_token": 1,  # 文本类型
    "uploader": 11,  # 人员
    "belonger": 11,  # 人员
    "type": 1,  # 文本类型
    "number": 1,  # 文本类型
    "date": 1,  # 文本类型
    "buyerName": 1,  # 文本类型
    "buyerTaxID": 1,  # 文本类型
    "sellerName": 1,  # 文本类型
    "sellerTaxID": 1,  # 文本类型
    "items_brief": 1,  # 文本类型
    "totalAmount": 2,  # 数字类型
    "error_message": 1,  # 文本类型
    "remark": 1,  # 文本类型
    "items": 1,  # 文本类型
    "item_num": 2,  # 数字类型
    "total_items_num": 2,  # 数字类型
    "items_unit": 1,  # 文本类型
    "status": 3,  # 状态类型
    # 状态类型: 0 - 待处理, -1 - 存在错误(解析错误/发票重复), -2 - 未通过自定义校验, 其余 - 自定义
}


def process_invoice_with_ocr(client, file_token: str, file_type: str,
                             base64_data: str, use_fallback: bool,
                             db: Database):
    import lark_oapi as lark
    import lark_oapi.api.drive.v1 as drive_v1
    client: lark.Client = client
    print(1)

    def perform_ocr(method: str):
        if "image" in file_type:
            return method("image", base64_data)
        elif "pdf" in file_type:
            return method("pdf", base64_data)
        else:
            raise ValueError("Unsupported file type")

    def check_duplicate(number: str):
        if not "invoices" in db.table_names():
            return None
        rows = list(db["invoices"].rows_where("number = ?", (number, )))
        for row in rows:
            if not row.get("error_message"):
                return row["file_token"]
        return None

    def insert_result(data: dict, processed: bool, error: str = None):
        record = {
            **data, "file_token": file_token,
            "processed": processed,
            "error_message": error,
            "status": '0' if error is None else '-1'
        }
        db["invoices"].insert(record,
                              pk="file_token",
                              replace=True,
                              alter=True)

    def get_file_tmp_download_url(file_token: str):
        request: drive_v1.BatchGetTmpDownloadUrlMediaRequest = drive_v1.BatchGetTmpDownloadUrlMediaRequest.builder() \
            .file_tokens(file_token) \
            .extra("请参考 [extra 参数说明](/ssl:ttdoc/uAjLw4CM/ukTMukTMukTM/reference/drive-v1/media/introduction)") \
            .build()

        # 发起请求
        response: drive_v1.BatchGetTmpDownloadUrlMediaResponse = client.drive.v1.media.batch_get_tmp_download_url(
            request)

        # 处理失败返回
        if not response.success():
            logger.error(
                f"client.drive.v1.media.batch_get_tmp_download_url failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}"
            )
            return None
        return response.data.tmp_download_urls[0].tmp_download_url

    try:
        ocr_result: Invoice = perform_ocr(BaiduOCR.vat_invoice_recognition)

        if not ocr_result.number or not ocr_result.totalAmount:
            raise ValueError("Missing required fields: number or totalAmount.")

        duplicate_token = check_duplicate(ocr_result.data.get('number'))
        if duplicate_token:
            msg = f"This file has been processed in file_token: {duplicate_token}"
            logger.debug(msg)
            insert_result(ocr_result.data, True, msg)
            return

        insert_result(ocr_result.data, True)
        logger.debug(f"Processed file {file_token} successfully.")

    except Exception as e:
        logger.warning(f"Primary OCR failed for file {file_token}: {e}")

        if use_fallback:
            try:
                ocr_result = perform_ocr(BaiduOCR.multiple_invoice_recognition)

                if not ocr_result.number or not ocr_result.totalAmount:
                    raise ValueError(
                        "Missing required fields in fallback OCR.")

                duplicate_token = check_duplicate(
                    ocr_result.data.get('number'))
                if duplicate_token:
                    msg = f"This file has been processed in file_token: {duplicate_token}"
                    logger.debug(msg)
                    insert_result(ocr_result.data, True, msg)
                    return

                insert_result(ocr_result.data, True)
                logger.debug(
                    f"Processed file {file_token} successfully (fallback).")

            except Exception as e:
                logger.error(
                    f"File {file_token} could not be processed with primary OCR and fallback: {e}.\nPlease check the file: {get_file_tmp_download_url(file_token)}"
                )
                error_msg = f"Fallback OCR failed: {e}"
                insert_result({}, False, error_msg)
        else:
            logger.error(
                f"File {file_token} could not be processed with primary OCR and fallback is disabled: {e}.\nPlease check the file: {get_file_tmp_download_url(file_token)}"
            )
            error_msg = f"This file cannot be processed: {e}"
            insert_result({}, False, error_msg)


def fetch_from_table(table_url: str,
                     db_path: str = "invoices.db",
                     use_fallback: bool = False):
    db = Database(db_path)
    lark_bitable_app_token, lark_bitable_table_id = extract_params_from_url(
        table_url)

    logger.info("Creating client for Lark API.")
    with yaspin(text="", spinner="dots") as spinner:
        # 延迟导入 lark_oapi，提高主程序启动速度
        import lark_oapi as lark
        import lark_oapi.api.drive.v1 as drive_v1
        import lark_oapi.api.bitable.v1 as bitable_v1

        if not (lark.APP_ID and lark.APP_SECRET):
            logger.error(
                "Lark APP_ID and APP_SECRET are not set. Please check file .env for LARK_APP_ID and LARK_APP_SECRET."
            )
            return
        client = lark.Client.builder() \
            .app_id(lark.APP_ID) \
            .app_secret(lark.APP_SECRET) \
            .log_level(LARK_LOG_LEVEL) \
            .build()
        spinner.ok("✅ Done")

    logger.info("Fetching records from the table...")
    with yaspin(text="", spinner="dots") as spinner:
        page_token = ""
        # update records table to ensure it is clean before inserting new records
        db['records'].delete_where("uid LIKE ?",
                                   (f"{lark_bitable_table_id}_%", ))
        while True:
            request: bitable_v1.SearchAppTableRecordRequest = bitable_v1.SearchAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .page_token(page_token) \
                .page_size(100) \
                .request_body(bitable_v1.SearchAppTableRecordRequestBody.builder()
                        .build()) \
                .build()

            response: bitable_v1.SearchAppTableRecordResponse = client.bitable.v1.app_table_record.search(
                request)

            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.search failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
                )
                return

            records = [{
                **record.fields, "uid":
                f"{lark_bitable_table_id}_{record.record_id}"
            } for record in response.data.items]
            lark.logger.debug(
                f"Fetched {len(records)} records from the table.")

            db["records"].insert_all(records,
                                     pk="uid",
                                     replace=True,
                                     alter=True)
            if not response.data.has_more:
                break
            page_token = response.data.page_token
        spinner.ok("✅ Done")

    logger.info("Processing invoice files...")
    if True:
        if not BaiduOCR.is_valid():
            logger.error(
                "Baidu OCR API credentials are not set or invalid. Please check file .env for BAIDU_OCR_API_KEY and BAIDU_OCR_SECRET_KEY."
            )
            return

        logger.info("Collecting invoices files...")
        result = db.execute(
            f"""
            SELECT
                records.uid,
                json_extract(value, '$.file_token') AS file_token,
                json_extract(value, '$.type') AS type
            FROM records, json_each(records.{INVOICE_COLUMN_NAME})
            WHERE records.uid LIKE ?
        """, (f"{lark_bitable_table_id}_%", )).fetchall()
        invoice_files = [{
            "record_uid": row[0],
            "file_token": row[1],
            "type": row[2],
        } for row in result]
        for invoice_file in tqdm(invoice_files, desc="Processing invoices"):
            if "invoices" in db.table_names():
                row = next(
                    db["invoices"].rows_where("file_token = ?",
                                              (invoice_file['file_token'], )),
                    None)
                if row and row.get("processed", False):
                    logger.debug(
                        f"File {invoice_file['file_token']} already processed, skipping."
                    )
                    continue

            request: drive_v1.DownloadMediaRequest = drive_v1.DownloadMediaRequest.builder() \
                .file_token(invoice_file['file_token']) \
                .build()

            response: drive_v1.DownloadMediaResponse = client.drive.v1.media.download(
                request)

            if not response.success():
                lark.logger.error(
                    f"client.drive.v1.media.download failed, code: {response.code}, msg: {response.msg},log_id: {response.get_log_id()}"
                )
                return

            # Read the file content and encode it to base64
            if response.file is not None:
                base64_data = base64.b64encode(
                    response.file.read()).decode("utf-8")
            else:
                invoice_file["error_message"] = "File is empty."
                logger.warning(
                    f"File {invoice_file['file_token']} is empty or not found."
                )

            process_invoice_with_ocr(client, invoice_file['file_token'],
                                     invoice_file['type'], base64_data,
                                     use_fallback, db)

    logger.info("Verifying invoice data with custom rules...")
    with yaspin(text="", spinner="dots") as spinner:
        invoices_data = db["invoices"].rows_where("processed = ?", (True, ))
        for invoice_data in tqdm(invoices_data, desc="Verifying invoices"):
            try:
                invoice = Invoice(invoice_data)
                verification_result = custom_rule.vertify_invoice(invoice)

                if verification_result["status"] == "error":
                    logger.debug(
                        f"Verification failed for file {invoice_data['file_token']}: {verification_result['message']}"
                    )
                    db["invoices"].update(
                        invoice_data['file_token'], {
                            "error_message": verification_result["message"],
                            "status": '-2'
                        })
                else:
                    logger.debug(
                        f"Verification passed for file {invoice_data['file_token']}."
                    )
            except Exception as e:
                logger.error(
                    f"Error verifying file {invoice_data['file_token']}: {str(e)}"
                )

        spinner.ok("✅ Done")

    logger.info("Updating records with invoice data...")
    with yaspin(text="", spinner="dots") as spinner:
        records_to_update = []
        result = db.execute("""
                SELECT 
                    invoices.file_token,
                    invoices.totalAmount,
                    invoices.error_message
                FROM invoices
            """).fetchall()
        invoices_by_token = {
            row[0]: {
                "total_amount": row[1],
                "error_message": row[2],
            }
            for row in result
        }

        records_in_current_table = list(db["records"].rows_where(
            "uid LIKE ?", (f"{lark_bitable_table_id}_%", )))
        for record in tqdm(records_in_current_table,
                           desc="Generating records to update."):
            result = db.execute(
                f"""
                SELECT 
                    json_extract(value, '$.file_token') AS file_token
                FROM records, json_each(records.{INVOICE_COLUMN_NAME})
                WHERE records.uid = ?
            """, (record['uid'], )).fetchall()
            invoice_files_token = [row[0] for row in result]
            error_message = ""
            total_amount = 0.0
            for index, invoice_file_token in enumerate(invoice_files_token,
                                                       start=1):
                if invoice_file_token in invoices_by_token:
                    invoice_data = invoices_by_token[invoice_file_token]
                    if invoice_data["error_message"]:
                        error_message += f"file index {{{index}}}: " + invoice_data[
                            "error_message"] + "; \n"
                    else:
                        total_amount += invoice_data["total_amount"]
            records_to_update.append({
                "record_id": record['uid'].split("_")[1],
                "fields": {
                    TOTAL_AMOUNT_COLUMN_NAME: float(total_amount),
                    APPROVAL_REMARKS_COLUMN_NAME: error_message,
                },
            })
        request: bitable_v1.BatchUpdateAppTableRecordRequest = bitable_v1.BatchUpdateAppTableRecordRequest.builder() \
            .app_token(lark_bitable_app_token) \
            .table_id(lark_bitable_table_id) \
            .request_body(bitable_v1.BatchUpdateAppTableRecordRequestBody.builder()
                .records([bitable_v1.AppTableRecord(update_data)
                    for update_data in records_to_update])
                .build()) \
            .build()
        response: bitable_v1.BatchUpdateAppTableRecordResponse = client.bitable.v1.app_table_record.batch_update(
            request)
        if not response.success():
            lark.logger.error(
                f"client.bitable.v1.app_table_record.batch_update failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
            )
            return
        logger.debug(f"Updating record {record['uid']} with invoice data.")
        logger.info(
            f"Updated {len(records_to_update)} records with invoice data in the table {lark_bitable_table_id}."
        )
        spinner.ok("✅ Done")
    logger.info(
        "All invoice files have been processed and the database has been updated."
    )


def export_to_local_document(db_path: str = "invoices.db",
                             output_path: str = "invoices.xlsx"):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import sqlite3
    from openpyxl.styles import numbers

    db = Database(db_path)
    logger.info("Exporting all tables to Excel...")

    table_names = db.table_names()
    if not table_names:
        logger.error("No tables found in the database.")
        return

    wb = Workbook()
    wb.remove(wb.active)

    # 使用 sqlite3 获取列类型
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    for table in table_names:
        try:
            # 获取列定义：名和类型
            cursor.execute(f"PRAGMA table_info({table})")
            col_info = {
                row["name"]: row["type"].lower()
                for row in cursor.fetchall()
            }

            # 获取数据
            df = pd.DataFrame(db[table].rows)
            ws = wb.create_sheet(title=table[:31])

            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # 设置格式
            for col_idx, column in enumerate(df.columns, 1):  # Excel 列索引从 1 开始
                col_type = col_info.get(column, "text")

                # 设置数字格式
                if "int" in col_type:
                    num_format = '0'
                elif "float" in col_type or "real" in col_type or "double" in col_type:
                    num_format = '0.00'
                elif "char" in col_type or "text" in col_type:
                    num_format = '@'
                elif "date" in col_type or "time" in col_type:
                    num_format = numbers.FORMAT_DATE_DATETIME
                else:
                    num_format = '@'  # 默认文本

                for row in ws.iter_rows(min_row=2,
                                        min_col=col_idx,
                                        max_col=col_idx):
                    for cell in row:
                        cell.number_format = num_format

            logger.info(f"Exported table '{table}' with {len(df)} rows.")
        except Exception as e:
            logger.warning(f"Failed to export table '{table}': {e}")

    wb.save(output_path)
    conn.close()
    logger.info(f"All tables exported to {output_path}.")


def create_lark_app_table(table_url: str, db_path: str = "invoices.db"):
    """
    (飞书)创建展示发票信息的数据表
    """
    db = Database(db_path)
    lark_bitable_app_token, _ = extract_params_from_url(table_url)

    logger.info("Creating client for Lark API.")
    with yaspin(text="", spinner="dots") as spinner:
        # 延迟导入 lark_oapi，提高主程序启动速度
        import lark_oapi as lark
        import lark_oapi.api.drive.v1 as drive_v1
        import lark_oapi.api.bitable.v1 as bitable_v1

        if not (lark.APP_ID and lark.APP_SECRET):
            logger.error(
                "Lark APP_ID and APP_SECRET are not set. Please check file .env for LARK_APP_ID and LARK_APP_SECRET."
            )
            return
        client = lark.Client.builder() \
            .app_id(lark.APP_ID) \
            .app_secret(lark.APP_SECRET) \
            .log_level(LARK_LOG_LEVEL) \
            .build()
        spinner.ok("✅ Done")

    logger.info("Creating app table.")
    with yaspin(text="", spinner="dots") as spinner:
        request: bitable_v1.CreateAppTableRequest = bitable_v1.CreateAppTableRequest.builder() \
            .app_token(lark_bitable_app_token) \
            .request_body(bitable_v1.CreateAppTableRequestBody.builder()
                .table(bitable_v1.ReqTable.builder()
                    .name("发票信息")
                    .default_view_name("全部信息")
                    .fields([
                        bitable_v1.AppTableCreateHeader.builder().field_name(field_name).type(type_value).property(bitable_v1.AppTableFieldProperty({"formatter": "0.00"})).build()
                            if type_value == 2 else (bitable_v1.AppTableCreateHeader.builder().field_name(field_name).type(type_value).build()
                             if type_value != 3 else
                             bitable_v1.AppTableCreateHeader.builder().field_name(field_name).type(type_value).property({"options": [
                                    {
                                        "name": "-2"
                                    },
                                    {
                                        "name": "-1"
                                    },
                                    {
                                        "name": "0"
                                    },
                                ]}).build())
                             for field_name, type_value in table_fields_type_map.items()
                        ])
                    .build())
                .build()) \
            .build()
        response: bitable_v1.CreateAppTableResponse = client.bitable.v1.app_table.create(
            request)
        if not response.success():
            if response.code == 1254013:
                logger.error('目标多维表格中已存在命名为“发票信息”的文档，创建行为失败.')
            else:
                lark.logger.error(
                    f"client.bitable.v1.app_table.create failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
                )
            return
        lark_bitable_table_id = response.data.table_id
        spinner.ok("✅ Done")

    logger.info("Inserting invoices data.")
    with yaspin(text="", spinner="dots") as spinner:
        # get invoices data
        keys = ("file_token", "type", "number", "date", "buyerTaxID",
                "buyerName", "sellerTaxID", "sellerName", "items_brief",
                "items_unit", "remark", "item_num", "total_items_num",
                "totalAmount", "error_message", "items", "status")
        result = db.execute(
            f"SELECT {', '.join(keys)} FROM invoices").fetchall()
        invoices_data = [{
            key: row[index]
            for index, key in enumerate(keys)
        } for row in result]

        # format all invoices data
        for invoice_data in invoices_data:
            for key, type_value in table_fields_type_map.items():
                if key in invoice_data:
                    value = invoice_data[key]

                    if value in (None, '', [], {}, ()):
                        del invoice_data[key]
                        continue

                    if type_value == 2:
                        invoice_data[key] = int(value)
                    elif type_value == 3:
                        invoice_data[key] = str(value)

        # add uploader and belonger data
        result = db.execute(f"""
            SELECT
                json_extract(records.{UPLOADER_COLUMN_NAME}, '$[0].id') AS uploader_id,
                json_extract(records.{BELONGER_COLUMN_NAME}, '$[0].id') AS belonger_id,
                json_extract(value, '$.file_token') AS file_token
            FROM records, json_each(records.{INVOICE_COLUMN_NAME})
        """).fetchall()

        invoices_by_token = {
            data['file_token']: data
            for data in invoices_data
        }
        for row in result:
            file_token = row[2]
            if file_token in invoices_by_token:
                invoice_data = invoices_by_token[file_token]
                if row[0]:
                    invoice_data['uploader'] = [{"id": row[0], "type": "user"}]
                if row[1]:
                    invoice_data['belonger'] = [{"id": row[1], "type": "user"}]

        records = [{"fields": data} for data in invoices_data]

        # try insert to bitable table
        BATCH_SIZE = 1000
        while records:
            batch = records[:BATCH_SIZE]
            request: bitable_v1.BatchCreateAppTableRecordRequest = bitable_v1.BatchCreateAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .ignore_consistency_check(True) \
                .request_body(bitable_v1.BatchCreateAppTableRecordRequestBody.builder()
                    .records(batch)
                    .build()) \
                .build()
            response: bitable_v1.BatchCreateAppTableRecordResponse = client.bitable.v1.app_table_record.batch_create(
                request)
            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.batch_create failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}"
                )
                return
            records = records[BATCH_SIZE:]
        spinner.ok("✅ Done")


def recheck_invoices(db_path: str = "invoices.db"):
    db = Database(db_path)
    logger.info("Verifying invoice data with custom rules...")
    with yaspin(text="", spinner="dots") as spinner:
        invoices_data = db["invoices"].rows_where("processed = ?", (True, ))
        for invoice_data in tqdm(invoices_data, desc="Verifying invoices"):
            try:
                invoice = Invoice(invoice_data)
                verification_result = custom_rule.vertify_invoice(invoice)

                if verification_result["status"] == "error":
                    logger.debug(
                        f"Verification failed for file {invoice_data['file_token']}: {verification_result['message']}"
                    )
                    db["invoices"].update(
                        invoice_data['file_token'], {
                            "error_message": verification_result["message"],
                            "status": '-2'
                        })
                else:
                    logger.debug(
                        f"Verification passed for file {invoice_data['file_token']}."
                    )
            except Exception as e:
                logger.error(
                    f"Error verifying file {invoice_data['file_token']}: {str(e)}"
                )

        spinner.ok("✅ Done")


def sync_from_table(table_url: str, db_path: str = "invoices.db"):
    """
    同步数据 方向 本地数据库<-远程云文档
    """
    db = Database(db_path)
    lark_bitable_app_token, lark_bitable_table_id = extract_params_from_url(
        table_url)

    logger.info("Creating client for Lark API.")
    with yaspin(text="", spinner="dots") as spinner:
        # 延迟导入 lark_oapi，提高主程序启动速度
        import lark_oapi as lark
        import lark_oapi.api.drive.v1 as drive_v1
        import lark_oapi.api.bitable.v1 as bitable_v1

        if not (lark.APP_ID and lark.APP_SECRET):
            logger.error(
                "Lark APP_ID and APP_SECRET are not set. Please check file .env for LARK_APP_ID and LARK_APP_SECRET."
            )
            return
        client = lark.Client.builder() \
            .app_id(lark.APP_ID) \
            .app_secret(lark.APP_SECRET) \
            .log_level(LARK_LOG_LEVEL) \
            .build()
        spinner.ok("✅ Done")

    logger.info("Fetching records from the table...")
    with yaspin(text="", spinner="dots") as spinner:
        page_token = ""
        # only need ["file_token", "status", "error_message"]
        while True:
            request: bitable_v1.SearchAppTableRecordRequest = bitable_v1.SearchAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .page_token(page_token) \
                .page_size(100) \
                .request_body(bitable_v1.SearchAppTableRecordRequestBody.builder()
                        .build()) \
                .build()

            response: bitable_v1.SearchAppTableRecordResponse = client.bitable.v1.app_table_record.search(
                request)

            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.search failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
                )
                return

            lark.logger.debug(
                f"Fetched {len(response.data.items)} records from the table.")

            for record in response.data.items:
                db["invoices"].update(
                    extract_text(record.fields, 'file_token'), {
                        "error_message":
                        extract_text(record.fields, 'error_message'),
                        "status":
                        extract_text(record.fields, 'status')
                    })
            if not response.data.has_more:
                break
            page_token = response.data.page_token
        spinner.ok("✅ Done")


def sync_to_table(table_url: str, db_path: str = "invoices.db"):
    """
    同步数据 方向 本地数据库->远程云文档
    """
    db = Database(db_path)
    lark_bitable_app_token, lark_bitable_table_id = extract_params_from_url(
        table_url)

    logger.info("Creating client for Lark API.")
    with yaspin(text="", spinner="dots") as spinner:
        # 延迟导入 lark_oapi，提高主程序启动速度
        import lark_oapi as lark
        import lark_oapi.api.drive.v1 as drive_v1
        import lark_oapi.api.bitable.v1 as bitable_v1

        if not (lark.APP_ID and lark.APP_SECRET):
            logger.error(
                "Lark APP_ID and APP_SECRET are not set. Please check file .env for LARK_APP_ID and LARK_APP_SECRET."
            )
            return
        client = lark.Client.builder() \
            .app_id(lark.APP_ID) \
            .app_secret(lark.APP_SECRET) \
            .log_level(LARK_LOG_LEVEL) \
            .build()
        spinner.ok("✅ Done")

    logger.info("Fetching table records.")
    with yaspin(text="", spinner="dots") as spinner:
        page_token = ""
        record_ids = {}
        while True:
            request: bitable_v1.SearchAppTableRecordRequest = bitable_v1.SearchAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .page_token(page_token) \
                .page_size(100) \
                .request_body(bitable_v1.SearchAppTableRecordRequestBody.builder()
                        .build()) \
                .build()

            response: bitable_v1.SearchAppTableRecordResponse = client.bitable.v1.app_table_record.search(
                request)

            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.search failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}"
                )
                return

            record_ids = record_ids | {
                extract_text(record.fields, 'file_token'): record.record_id
                for record in response.data.items
            }
            lark.logger.debug(
                f"Fetched {len(record_ids)} record_ids from the table.")

            if not response.data.has_more:
                break
            page_token = response.data.page_token
        spinner.ok("✅ Done")

    logger.info("Updating invoices data.")
    with yaspin(text="", spinner="dots") as spinner:
        # get invoices data
        keys = ("file_token", "type", "number", "date", "buyerTaxID",
                "buyerName", "sellerTaxID", "sellerName", "items_brief",
                "items_unit", "remark", "item_num", "total_items_num",
                "totalAmount", "error_message", "items", "status")
        result = db.execute(
            f"SELECT {', '.join(keys)} FROM invoices").fetchall()
        invoices_data = [{
            key: row[index]
            for index, key in enumerate(keys)
        } for row in result]

        # format all invoices data
        for invoice_data in invoices_data:
            for key, type_value in table_fields_type_map.items():
                if key in invoice_data:
                    value = invoice_data[key]

                    if value in (None, '', [], {}, ()):
                        del invoice_data[key]
                        continue

                    if type_value == 2:
                        invoice_data[key] = int(value)
                    elif type_value == 3:
                        invoice_data[key] = str(value)

        # add uploader and belonger data
        result = db.execute(f"""
            SELECT
                json_extract(records.{UPLOADER_COLUMN_NAME}, '$[0].id') AS uploader_id,
                json_extract(records.{BELONGER_COLUMN_NAME}, '$[0].id') AS belonger_id,
                json_extract(value, '$.file_token') AS file_token
            FROM records, json_each(records.{INVOICE_COLUMN_NAME})
        """).fetchall()

        invoices_by_token = {
            data['file_token']: data
            for data in invoices_data
        }
        for row in result:
            file_token = row[2]
            if file_token in invoices_by_token:
                invoice_data = invoices_by_token[file_token]
                if row[0]:
                    invoice_data['uploader'] = [{"id": row[0], "type": "user"}]
                if row[1]:
                    invoice_data['belonger'] = [{"id": row[1], "type": "user"}]

        update_records = []
        insert_records = []
        for data in invoices_data:
            if data['file_token'] in record_ids:
                update_records.append({
                    "fields": {
                        "error_message": data.get('error_message'),
                        "status": data.get('status'),
                    },
                    "record_id":
                    record_ids[data['file_token']]
                })
            else:
                insert_records.append({"fields": data})

        # try upate to bitable table
        BATCH_SIZE = 1000
        while update_records:
            batch = update_records[:BATCH_SIZE]
            request: bitable_v1.BatchUpdateAppTableRecordRequest = bitable_v1.BatchUpdateAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .request_body(bitable_v1.BatchUpdateAppTableRecordRequestBody.builder()
                    .records(batch)
                    .build()) \
                .build()
            response: bitable_v1.BatchUpdateAppTableRecordResponse = client.bitable.v1.app_table_record.batch_update(
                request)
            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.batch_update failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}"
                )
                return
            update_records = update_records[BATCH_SIZE:]

        # Insert invoice records that do not exist in the table
        while insert_records:
            batch = insert_records[:BATCH_SIZE]
            request: bitable_v1.BatchCreateAppTableRecordRequest = bitable_v1.BatchCreateAppTableRecordRequest.builder() \
                .app_token(lark_bitable_app_token) \
                .table_id(lark_bitable_table_id) \
                .ignore_consistency_check(True) \
                .request_body(bitable_v1.BatchCreateAppTableRecordRequestBody.builder()
                    .records(batch)
                    .build()) \
                .build()
            response: bitable_v1.BatchCreateAppTableRecordResponse = client.bitable.v1.app_table_record.batch_create(
                request)
            if not response.success():
                lark.logger.error(
                    f"client.bitable.v1.app_table_record.batch_create failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}"
                )
                return
            insert_records = insert_records[BATCH_SIZE:]
        spinner.ok("✅ Done")


def auto_sync(table_url: str, db_path: str = "invoices.db"):
    """
    根据修改时间自动选择同步方向
    """
    db = Database(db_path)
    lark_bitable_app_token, lark_bitable_table_id = extract_params_from_url(
        table_url)

    logger.info("Creating client for Lark API.")
    with yaspin(text="", spinner="dots") as spinner:
        # 延迟导入 lark_oapi，提高主程序启动速度
        import lark_oapi as lark
        import lark_oapi.api.drive.v1 as drive_v1
        import lark_oapi.api.bitable.v1 as bitable_v1

        if not (lark.APP_ID and lark.APP_SECRET):
            logger.error(
                "Lark APP_ID and APP_SECRET are not set. Please check file .env for LARK_APP_ID and LARK_APP_SECRET."
            )
            return
        client = lark.Client.builder() \
            .app_id(lark.APP_ID) \
            .app_secret(lark.APP_SECRET) \
            .log_level(LARK_LOG_LEVEL) \
            .build()
        spinner.ok("✅ Done")

    logger.info(
        "Comparing local and remote revision to determine sync direction.")
    remote_revision = 0
    local_record_revision = 0
    result = ""
    with yaspin(text="", spinner="dots") as spinner:
        # fetch remote revision
        def fetch_remote_revision():
            remote_revision = ""
            page_token = ""
            while remote_revision == "":
                request: bitable_v1.ListAppTableRequest = bitable_v1.ListAppTableRequest.builder() \
                    .app_token(lark_bitable_app_token) \
                    .page_token(page_token) \
                    .page_size(100) \
                    .build()
                response: bitable_v1.ListAppTableResponse = client.bitable.v1.app_table.list(
                    request)
                if not response.success():
                    lark.logger.error(
                        f"client.bitable.v1.app_table.list failed, code: {response.code}, msg: {response.msg}, log_id: {response.get_log_id()}, resp: \n{json.dumps(json.loads(response.raw.content), indent=4, ensure_ascii=False)}"
                    )
                    return 0

                table_info = next(
                    (table_info for table_info in response.data.items
                     if table_info.table_id == lark_bitable_table_id), None)
                if table_info:
                    return table_info.revision
                if not response.data.has_more:
                    logger.exception('cannot fetch remote table revision.')
                page_token = response.data.page_token

        remote_revision = fetch_remote_revision()
        logger.debug(f'remote_revision is {remote_revision}')
        # fetch local revision
        try:
            local_record_revision = db['remote_table_revision'].get(
                lark_bitable_table_id).get('value')
            if not local_record_revision:
                local_record_revision = 0
        except Exception:
            db['remote_table_revision'].insert(
                {
                    "table_id": lark_bitable_table_id,
                    "value": 0,
                },
                pk="table_id")
            local_record_revision = 0
        logger.debug(f'local_record_revision is {local_record_revision}')

        # compare local and remote revision
        if local_record_revision >= remote_revision:
            result = "sync_to_table"
        else:
            result = "sync_from_table"

        spinner.ok("✅ Done")

    # write current revision to database
    if result == "sync_to_table":
        logger.info('Select sync_to_table.')
        sync_to_table(table_url, db_path)
        remote_revision = fetch_remote_revision()
        db['remote_table_revision'].update(lark_bitable_table_id,
                                           {"value": remote_revision})
    elif result == "sync_from_table":
        logger.info('Select sync_from_table.')
        sync_from_table(table_url, db_path)
        db['remote_table_revision'].update(lark_bitable_table_id,
                                           {"value": remote_revision})


def group_invoices(file_path, db_path: str = "invoices.db"):
    db = Database(db_path)

    logger.info('Parsing target file.')
    with yaspin(text="", spinner="dots") as spinner:
        with open(file_path, 'r') as f:
            request = json.loads(f.read())
        try:
            key: str = request['key']
            target: dict = request['target']
            if not isinstance(key, str) or not isinstance(target, dict):
                raise ValueError('key or target has wrong type.')
        except Exception:
            logger.exception("传入的参数文件格式错误，请检查 JSON 格式或字段内容")

        spinner.ok("✅ Done")

    logger.info('Updating invoices data.')
    with yaspin(text="", spinner="dots") as spinner:
        for key_word, status in target.items():
            db.execute(f"UPDATE invoices SET status = ? WHERE {key} = ?",
                       (status, key_word))
        db.conn.commit()
        spinner.ok("✅ Done")
